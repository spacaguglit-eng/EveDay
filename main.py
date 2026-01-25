import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import json
import os
import threading
import time
import subprocess
from copy import copy
from functools import wraps
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import calendar
import openpyxl
from history_manager import HistoryManager

# Для быстрого копирования листов через Excel COM
try:
    import win32com.client
    import pythoncom
    HAS_WIN32COM = True
except ImportError:
    pythoncom = None
    HAS_WIN32COM = False

# --- КОНФИГУРАЦИЯ ---
CONFIG_FILE = "app_config.json"
DEFAULT_CONFIG = {
    "file_paths": [""] * 11,
    "min_downtime": 10,
    "excluded_categories": "Обед, Перерыв"
}
MONTHS = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

class ConfigManager:
    @staticmethod
    def load_config():
        data = DEFAULT_CONFIG.copy()
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    data.update(loaded)
                    paths = data.get("file_paths", [])
                    if len(paths) < 11:
                        paths.extend([""] * (11 - len(paths)))
                    data["file_paths"] = paths[:11]
            except Exception as e:
                print(f"Ошибка загрузки конфига: {e}")
        return data

    @staticmethod
    def save_config(paths, min_downtime, excluded_categories):
        try:
            data = {
                "file_paths": paths,
                "min_downtime": min_downtime,
                "excluded_categories": excluded_categories
            }
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Ошибка сохранения конфига: {e}")

# --- СТРУКТУРЫ ДАННЫХ ---
class ProblemEntry:
    def __init__(self, filename, sheet, shift, time_val, type_val, formulation, comment=""):
        self.filename = filename
        self.shift = shift
        self.time_val = time_val
        self.type_val = str(type_val).strip() if type_val else "Не указан"
        self.formulation = self._clean_text(formulation)
        self.comment = self._clean_text(comment)

    def _clean_text(self, text):
        if not text: return ""
        s = " ".join(str(text).split())
        return s[0].upper() + s[1:] if s else ""

class LineData:
    """Хранит информацию о линии для отчета и сборки файла"""
    def __init__(self, filepath, sheet_name, line_name, plan, fact, problems):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.line_name = line_name
        self.plan = plan
        self.fact = fact
        self.problems = problems

# --- УТИЛИТЫ EXCEL ---
class ExcelComContext:
    def __init__(self):
        self.excel = None

    def __enter__(self):
        if pythoncom:
            pythoncom.CoInitialize()
        self.excel = win32com.client.DispatchEx("Excel.Application")
        self.excel.Visible = False
        self.excel.DisplayAlerts = False
        self.excel.ScreenUpdating = False
        return self.excel

    def __exit__(self, exc_type, exc, tb):
        had_error = exc_type is not None
        self._safe_close()
        if had_error:
            self._force_kill()
        if pythoncom:
            pythoncom.CoUninitialize()
        return False

    def _safe_close(self):
        if not self.excel:
            return
        try:
            for wb in list(self.excel.Workbooks):
                try:
                    wb.Close(False)
                except:
                    pass
        except:
            pass
        try:
            self.excel.Quit()
        except:
            pass
        self.excel = None

    def _force_kill(self):
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        except:
            pass

def retry_on_failure(max_attempts=3, base_delay=1):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            log_callback = kwargs.get("log_callback")
            for attempt in range(1, max_attempts + 1):
                if log_callback:
                    log_callback(f"Запуск попытки {attempt}/{max_attempts}")
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if log_callback:
                        log_callback(f"Попытка {attempt}/{max_attempts} не удалась: {e}")
                    if attempt >= max_attempts:
                        raise
                    time.sleep(base_delay * (2 ** (attempt - 1)))
        return wrapper
    return decorator

def _sanitize_sheet_name(name):
    safe_name = name[:31]
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        safe_name = safe_name.replace(ch, '_')
    return safe_name

@retry_on_failure(max_attempts=3, base_delay=1)
def copy_sheets_fast(lines_data, save_path, progress_callback=None, log_callback=None):
    """
    Быстрое копирование листов через Excel COM (win32com).
    Копирует листы целиком со всеми стилями за секунды.
    
    progress_callback: функция(value) где value от 0 до 100
    """
    if not HAS_WIN32COM:
        raise ImportError("Требуется pywin32: pip install pywin32")
    
    # Нормализуем путь для Windows (абсолютный путь с обратными слэшами)
    save_path = os.path.abspath(save_path)
    
    # Если файл существует - удаляем его
    if os.path.exists(save_path):
        try:
            os.remove(save_path)
        except PermissionError:
            raise PermissionError(f"Файл {save_path} открыт в другой программе. Закройте его и попробуйте снова.")
    
    total = len(lines_data)

    if progress_callback:
        progress_callback(5)

    with ExcelComContext() as excel:
        new_wb = excel.Workbooks.Add()
        default_sheets = [new_wb.Worksheets(i).Name for i in range(1, new_wb.Worksheets.Count + 1)]

        for idx, ld in enumerate(lines_data):
            src_wb = None
            try:
                src_path = os.path.abspath(ld.filepath)
                src_wb = excel.Workbooks.Open(src_path, ReadOnly=True)
                src_ws = src_wb.Worksheets(ld.sheet_name)
                src_ws.Copy(After=new_wb.Worksheets(new_wb.Worksheets.Count))

                copied_sheet = new_wb.Worksheets(new_wb.Worksheets.Count)
                copied_sheet.Name = _sanitize_sheet_name(ld.line_name)
            except Exception as e:
                if log_callback:
                    log_callback(f"Ошибка копирования {ld.line_name}: {e}")
                else:
                    print(f"Ошибка копирования {ld.line_name}: {e}")
            finally:
                if src_wb:
                    src_wb.Close(False)

            if progress_callback and total > 0:
                progress_callback(10 + int(80 * (idx + 1) / total))
            time.sleep(0)

        for sheet_name in default_sheets:
            try:
                new_wb.Worksheets(sheet_name).Delete()
            except:
                pass

        if progress_callback:
            progress_callback(95)

        sheets_count = new_wb.Worksheets.Count
        new_wb.SaveAs(save_path, FileFormat=51)
        new_wb.Close(False)

        if progress_callback:
            progress_callback(100)

        return sheets_count

def copy_sheets_openpyxl(lines_data, save_path, progress_callback=None):
    if progress_callback:
        progress_callback(5)

    if os.path.exists(save_path):
        try:
            os.remove(save_path)
        except PermissionError:
            raise PermissionError(f"Файл {save_path} открыт в другой программе. Закройте его и попробуйте снова.")

    new_wb = openpyxl.Workbook()
    if new_wb.active:
        new_wb.remove(new_wb.active)

    total = len(lines_data)
    for idx, ld in enumerate(lines_data):
        src_wb = None
        try:
            src_path = os.path.abspath(ld.filepath)
            src_wb = openpyxl.load_workbook(src_path, data_only=False)
            if ld.sheet_name not in src_wb.sheetnames:
                continue
            src_ws = src_wb[ld.sheet_name]

            new_ws = new_wb.create_sheet(title=_sanitize_sheet_name(ld.line_name))

            for col, dim in src_ws.column_dimensions.items():
                new_ws.column_dimensions[col].width = dim.width
            for row, dim in src_ws.row_dimensions.items():
                new_ws.row_dimensions[row].height = dim.height

            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
        finally:
            if src_wb:
                src_wb.close()

        if progress_callback and total > 0:
            progress_callback(10 + int(80 * (idx + 1) / total))
        time.sleep(0)

    if progress_callback:
        progress_callback(95)

    sheets_count = len(new_wb.sheetnames)
    new_wb.save(save_path)

    if progress_callback:
        progress_callback(100)

    return sheets_count

def copy_sheets_with_fallback(lines_data, save_path, progress_callback=None, log_callback=None):
    try:
        sheets_count = copy_sheets_fast(lines_data, save_path, progress_callback=progress_callback, log_callback=log_callback)
        return sheets_count, "COM"
    except Exception as e:
        if log_callback:
            log_callback(f"COM не сработал, переключаюсь на openpyxl: {e}")
        sheets_count = copy_sheets_openpyxl(lines_data, save_path, progress_callback=progress_callback)
        return sheets_count, "openpyxl"

def build_summary_text(lines_data):
    text_lines = []
    text_lines.append("Выработка по линиям:")
    text_lines.append("")

    if not lines_data:
        text_lines.append("Нет данных (все листы пустые или файлы недоступны).")
    else:
        for ld in lines_data:
            fact_fmt = "{:,.0f}".format(ld.fact).replace(',', ' ')
            plan_fmt = "{:,.0f}".format(ld.plan).replace(',', ' ')

            text_lines.append(f"{ld.line_name} - {fact_fmt} шт. (План: {plan_fmt})")

            if ld.problems:
                for p in ld.problems:
                    reason = p.formulation if p.formulation else p.type_val
                    comment_part = f" | {p.comment}" if p.comment else ""
                    text_lines.append(f"  ⚠️ {reason} ({p.time_val:g} мин){comment_part}")

            text_lines.append("")

    return "\n".join(text_lines)

# --- ЛОГИКА ОБРАБОТКИ ---
class ExcelProcessor:
    def __init__(self, log_callback, progress_callback, status_callback=None, cancel_event=None):
        self.log = log_callback
        self.update_progress = progress_callback
        self.status_callback = status_callback
        self.cancel_event = cancel_event
        self.is_running = False

    def process_sheet(self, sheet, sheet_name, filename, min_minutes_threshold, excluded_list):
        # 1. Считываем данные в кэш
        MIN_ROW, MAX_ROW = 21, 205
        data_cache = {}
        try:
            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=1, max_col=13, values_only=True), start=MIN_ROW):
                keep = False
                if (21 <= row_idx <= 42) or (136 <= row_idx <= 158): keep = True
                elif (47 <= row_idx <= 113) or (162 <= row_idx <= 205): keep = True
                if keep:
                    data_cache[row_idx] = list(row_values)
        except Exception as e:
            return None, [], 0, 0, f"Ошибка чтения: {e}"

        # 2. Проверка: Пустой ли лист?
        def is_range_empty(rows):
            for r in rows:
                row_data = data_cache.get(r)
                val = row_data[0] if row_data and len(row_data) > 0 else None
                if val is not None and str(val).strip() != "": return False
            return True

        # Проверяем диапазоны смен (упрощенно по ключевым ячейкам операторов)
        if is_range_empty(range(37, 43)) and is_range_empty(range(152, 158)):
            return None, [], 0, 0, "Лист пустой"

        # 3. Сбор статистики (План/Факт)
        def sum_vals(rows, col_idx):
            total = 0.0
            for r in rows:
                row_data = data_cache.get(r)
                if row_data and len(row_data) > col_idx:
                    val = row_data[col_idx]
                    if val:
                        if isinstance(val, str): val = val.replace(',', '.').replace(' ', '')
                        try: total += float(val)
                        except: pass
            return total

        day_plan = sum_vals(range(21, 33, 2), 9)
        day_fact = sum_vals(range(21, 33, 2), 10)
        night_plan = sum_vals(range(136, 148, 2), 9)
        night_fact = sum_vals(range(136, 148, 2), 10)
        
        total_plan = day_plan + night_plan
        total_fact = day_fact + night_fact

        # 4. Поиск проблем
        def get_problems(rows, shift):
            res = []
            for r in rows:
                row_data = data_cache.get(r)
                if not row_data: continue
                try:
                    time_v = row_data[10]
                    if not time_v: continue
                    if isinstance(time_v, str): time_v = float(time_v.replace(',', '.').replace(' ', ''))
                    else: time_v = float(time_v)
                    
                    if time_v < min_minutes_threshold: continue
                    
                    type_v = str(row_data[7]).strip() if row_data[7] else ""
                    
                    is_excluded = any(ex.lower() in type_v.lower() for ex in excluded_list)
                    if is_excluded: continue

                    # Комментарий из столбца L (индекс 11)
                    comment_v = row_data[11] if len(row_data) > 11 else ""
                    
                    res.append(ProblemEntry(filename, sheet_name, shift, time_v, type_v, row_data[5], comment_v))
                except: pass
            return res

        problems = []
        problems.extend(get_problems(range(47, 114), "ДЕНЬ"))
        problems.extend(get_problems(range(162, 206), "НОЧЬ"))
        problems.sort(key=lambda x: x.time_val, reverse=True)
        
        return True, problems[:2], total_plan, total_fact, "OK"

    def run(self, file_paths, target_day, target_month_str, target_year, min_downtime, exclude_str):
        self.is_running = True
        excluded_list = [x.strip() for x in exclude_str.split(',') if x.strip()]
        
        valid_lines_data = [] # Список объектов LineData для непустых листов
        
        valid_paths = [p for p in file_paths if p.strip()]
        total_files = len(valid_paths)
        sheet_name = str(target_day)
        processed_count = 0
        lock = threading.Lock()

        def update_line(line_name, progress, status, message):
            if self.status_callback:
                self.status_callback(line_name, progress, status, message)

        def process_one(file_path):
            if self.cancel_event and self.cancel_event.is_set():
                file_path = os.path.normpath(file_path)
                fname = os.path.basename(file_path)
                line_name = os.path.splitext(fname)[0]
                update_line(line_name, 100, "ошибка", f"{line_name}: Отменено")
                return None

            file_path = os.path.normpath(file_path)
            fname = os.path.basename(file_path)
            line_name = os.path.splitext(fname)[0]

            update_line(line_name, 10, "обработка", f"{line_name}: Открытие файла... 10%")
            self.log(f"Проверка: {line_name}...")

            if not os.path.exists(file_path):
                update_line(line_name, 100, "ошибка", f"{line_name}: Файл не найден")
                self.log(f"  Ошибка: Файл не найден.")
                return None

            try:
                update_line(line_name, 30, "обработка", f"{line_name}: Чтение данных... 30%")
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if self.cancel_event and self.cancel_event.is_set():
                    update_line(line_name, 100, "ошибка", f"{line_name}: Отменено")
                    wb.close()
                    return None
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    update_line(line_name, 70, "обработка", f"{line_name}: Анализ проблем... 70%")
                    is_valid, probs, plan, fact, msg = self.process_sheet(ws, sheet_name, fname, min_downtime, excluded_list)

                    if is_valid:
                        line_obj = LineData(file_path, sheet_name, line_name, plan, fact, probs)
                        with lock:
                            valid_lines_data.append(line_obj)
                        update_line(line_name, 100, "готово", f"{line_name}: ✓ Готово 100%")
                        self.log(f"  OK. Факт: {fact:g}")
                    else:
                        update_line(line_name, 100, "ошибка", f"{line_name}: {msg}")
                        self.log(f"  Пропуск: {msg}")
                else:
                    update_line(line_name, 100, "ошибка", f"{line_name}: Лист {sheet_name} не найден")
                    self.log(f"  Пропуск: Лист {sheet_name} не найден")
                wb.close()
            except Exception as e:
                update_line(line_name, 100, "ошибка", f"{line_name}: Ошибка {e}")
                self.log(f"  Ошибка: {e}")

            return True

        if total_files == 0:
            self.is_running = False
            self.update_progress(0)
            return valid_lines_data

        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = []
            for path in valid_paths:
                if self.cancel_event and self.cancel_event.is_set():
                    break
                futures.append(executor.submit(process_one, path))

            for future in as_completed(futures):
                if self.cancel_event and self.cancel_event.is_set():
                    for f in futures:
                        f.cancel()
                    break
                try:
                    future.result()
                except Exception as e:
                    self.log(f"  Ошибка потока: {e}")
                with lock:
                    processed_count += 1
                    self.update_progress((processed_count / total_files) * 100)

        self.is_running = False
        return valid_lines_data

# --- GUI: СЕТКА СТАТУСОВ ---
class StatusGrid(ttk.Frame):
    def __init__(self, parent, rows=11):
        super().__init__(parent)
        self.rows = rows
        self.items = {}
        self._build()

    def _build(self):
        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text="Линия", width=20).grid(row=0, column=0, sticky="w")
        ttk.Label(header, text="Прогресс", width=20).grid(row=0, column=1, sticky="w", padx=5)
        ttk.Label(header, text="Статус").grid(row=0, column=2, sticky="w")

        for i in range(self.rows):
            row_frame = ttk.Frame(self)
            row_frame.grid(row=i + 1, column=0, sticky="ew", pady=1)

            name_label = ttk.Label(row_frame, text=f"Линия {i+1}", width=20)
            name_label.grid(row=0, column=0, sticky="w")

            prog = ttk.Progressbar(row_frame, mode="determinate", length=160)
            prog.grid(row=0, column=1, sticky="w", padx=5)

            status_label = tk.Label(row_frame, text="ожидание", fg="gray")
            status_label.grid(row=0, column=2, sticky="w")

            self.items[i] = {
                "name": name_label,
                "progress": prog,
                "status": status_label
            }

    def set_lines(self, line_names):
        for i in range(self.rows):
            name = line_names[i] if i < len(line_names) else f"Линия {i+1}"
            self.items[i]["name"].config(text=name)
            self.items[i]["progress"]["value"] = 0
            self.items[i]["status"].config(text="ожидание", fg="gray")

    def update_line(self, line_name, progress, status, message):
        color_map = {
            "ожидание": "gray",
            "обработка": "blue",
            "готово": "green",
            "ошибка": "red"
        }
        for i in range(self.rows):
            if self.items[i]["name"].cget("text") == line_name:
                self.items[i]["progress"]["value"] = progress
                self.items[i]["status"].config(text=message, fg=color_map.get(status, "gray"))
                break

# --- GUI: РЕДАКТОР И СВОДКА ---
class SummaryFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.lines_data = []
        
        # Верхняя часть: редактор проблем
        edit_frame = ttk.LabelFrame(self, text="Редактирование проблем/простоев", padding=5)
        edit_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Таблица проблем
        columns = ("line", "time", "type", "description", "comment")
        self.tree = ttk.Treeview(edit_frame, columns=columns, show="headings", height=10)
        self.tree.heading("line", text="Линия")
        self.tree.heading("time", text="Время (мин)")
        self.tree.heading("type", text="Тип")
        self.tree.heading("description", text="Описание")
        self.tree.heading("comment", text="Комментарий")
        self.tree.column("line", width=100)
        self.tree.column("time", width=70)
        self.tree.column("type", width=90)
        self.tree.column("description", width=200)
        self.tree.column("comment", width=200)
        
        tree_scroll = ttk.Scrollbar(edit_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        
        # Двойной клик для редактирования
        self.tree.bind("<Double-1>", self._on_tree_double_click)
        
        # Кнопки редактирования
        edit_btn_frame = ttk.Frame(self)
        edit_btn_frame.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Button(edit_btn_frame, text="✏️ Редактировать", command=self._edit_selected).pack(side="left", padx=2)
        ttk.Button(edit_btn_frame, text="➕ Добавить", command=self._add_problem).pack(side="left", padx=2)
        ttk.Button(edit_btn_frame, text="🗑️ Удалить", command=self._delete_selected).pack(side="left", padx=2)
        
        # Нижняя часть: текстовая сводка
        self.txt_area = scrolledtext.ScrolledText(self, font=("Consolas", 10), height=8)
        self.txt_area.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Прогресс-бар для сохранения
        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", padx=5, pady=(0, 5))
        self.progress.pack_forget()  # Скрываем до использования
        
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Button(btn_frame, text="🔄 Обновить сводку", command=self._refresh_summary).pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_copy = ttk.Button(btn_frame, text="📋 Копировать текст", command=self.copy_text)
        self.btn_copy.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_save_history = ttk.Button(btn_frame, text="💾 Сохранить в Историю", command=self.save_to_history)
        self.btn_save_history.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        self.btn_excel = ttk.Button(btn_frame, text="💾 СОЗДАТЬ СВОДНЫЙ EXCEL ФАЙЛ", command=self.create_consolidated_excel)
        self.btn_excel.pack(side="right", fill="x", expand=True, padx=(5, 0))

    def populate(self, lines_data, prebuilt_text=None):
        self.lines_data = lines_data
        self._populate_tree()
        self._refresh_summary()

    def _populate_tree(self):
        # Очищаем таблицу
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Заполняем проблемами
        for ld in self.lines_data:
            for p in ld.problems:
                self.tree.insert("", "end", values=(
                    ld.line_name,
                    f"{p.time_val:g}",
                    p.type_val,
                    p.formulation,
                    p.comment
                ), tags=(ld.line_name,))

    def _refresh_summary(self):
        final_text = build_summary_text(self.lines_data)
        self.txt_area.config(state='normal')
        self.txt_area.delete("1.0", tk.END)
        self.txt_area.insert("1.0", final_text)
        self.txt_area.config(state='disabled')

    def _on_tree_double_click(self, event):
        self._edit_selected()

    def _edit_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Выберите", "Выберите проблему для редактирования")
            return
        
        item = selected[0]
        values = self.tree.item(item, "values")
        line_name, time_val, type_val, desc, comment = values
        
        # Типы простоев для выбора
        type_options = ["общезаводские", "технологические", "организационные", "механические", "электрические", "прочие"]
        
        # Современное окно редактирования
        edit_win = tk.Toplevel(self)
        edit_win.title("Редактирование проблемы")
        edit_win.geometry("550x380")
        edit_win.resizable(False, False)
        edit_win.transient(self)
        edit_win.grab_set()
        
        # Центрирование окна
        edit_win.update_idletasks()
        x = (edit_win.winfo_screenwidth() - 550) // 2
        y = (edit_win.winfo_screenheight() - 380) // 2
        edit_win.geometry(f"550x380+{x}+{y}")
        
        # Основной контейнер с отступами
        main_frame = ttk.Frame(edit_win, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # --- Линия (read-only) ---
        line_frame = ttk.Frame(main_frame)
        line_frame.pack(fill="x", pady=(0, 12))
        ttk.Label(line_frame, text="Линия:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        line_label = ttk.Label(line_frame, text=line_name, foreground="#555", font=("Segoe UI", 9))
        line_label.pack(anchor="w", pady=(2, 0))
        
        # --- Время и Тип в одной строке ---
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill="x", pady=(0, 12))
        
        # Время
        time_frame = ttk.Frame(row_frame)
        time_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Label(time_frame, text="Время простоя:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        time_input_frame = ttk.Frame(time_frame)
        time_input_frame.pack(anchor="w", pady=(2, 0))
        time_var = tk.StringVar(value=time_val)
        time_entry = ttk.Entry(time_input_frame, textvariable=time_var, width=10, font=("Segoe UI", 10))
        time_entry.pack(side="left")
        ttk.Label(time_input_frame, text="мин", foreground="#666").pack(side="left", padx=(5, 0))
        
        # Тип (Combobox)
        type_frame = ttk.Frame(row_frame)
        type_frame.pack(side="left", fill="x", expand=True)
        ttk.Label(type_frame, text="Тип:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        type_var = tk.StringVar(value=type_val)
        type_combo = ttk.Combobox(type_frame, textvariable=type_var, values=type_options, width=20, font=("Segoe UI", 10))
        type_combo.pack(anchor="w", pady=(2, 0))
        
        # --- Описание ---
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(fill="x", pady=(0, 12))
        ttk.Label(desc_frame, text="Описание:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        desc_var = tk.StringVar(value=desc)
        desc_entry = ttk.Entry(desc_frame, textvariable=desc_var, font=("Segoe UI", 10))
        desc_entry.pack(fill="x", pady=(2, 0))
        
        # --- Комментарий (многострочный) ---
        comment_frame = ttk.Frame(main_frame)
        comment_frame.pack(fill="both", expand=True, pady=(0, 15))
        ttk.Label(comment_frame, text="Комментарий:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        comment_text = tk.Text(comment_frame, height=4, font=("Segoe UI", 10), wrap="word", relief="solid", borderwidth=1)
        comment_text.pack(fill="both", expand=True, pady=(2, 0))
        comment_text.insert("1.0", comment if comment else "")
        
        # --- Кнопки ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x")
        
        def save_changes():
            new_time = time_var.get()
            new_type = type_var.get()
            new_desc = desc_var.get()
            new_comment = comment_text.get("1.0", "end-1c").strip()
            
            # Валидация времени
            try:
                float(new_time.replace(",", "."))
            except:
                messagebox.showwarning("Ошибка", "Введите корректное время (число)")
                return
            
            # Обновляем в таблице
            self.tree.item(item, values=(line_name, new_time, new_type, new_desc, new_comment))
            
            # Обновляем в данных
            for ld in self.lines_data:
                if ld.line_name == line_name:
                    for p in ld.problems:
                        if str(p.time_val) == time_val.replace(",", ".") or f"{p.time_val:g}" == time_val:
                            try:
                                p.time_val = float(new_time.replace(",", "."))
                            except:
                                pass
                            p.type_val = new_type
                            p.formulation = new_desc
                            p.comment = new_comment
                            break
                    break
            
            self._refresh_summary()
            edit_win.destroy()
        
        ttk.Button(btn_frame, text="Отмена", command=edit_win.destroy, width=12).pack(side="right", padx=(5, 0))
        
        style = ttk.Style()
        style.configure("Primary.TButton", font=("Segoe UI", 9, "bold"))
        ttk.Button(btn_frame, text="Сохранить", command=save_changes, width=12, style="Primary.TButton").pack(side="right")

    def _add_problem(self):
        if not self.lines_data:
            messagebox.showwarning("Пусто", "Сначала запустите анализ")
            return
        
        # Типы простоев для выбора
        type_options = ["общезаводские", "технологические", "организационные", "механические", "электрические", "прочие"]
        line_names = [ld.line_name for ld in self.lines_data]
        
        # Современное окно добавления
        add_win = tk.Toplevel(self)
        add_win.title("Добавить проблему")
        add_win.geometry("550x400")
        add_win.resizable(False, False)
        add_win.transient(self)
        add_win.grab_set()
        
        # Центрирование окна
        add_win.update_idletasks()
        x = (add_win.winfo_screenwidth() - 550) // 2
        y = (add_win.winfo_screenheight() - 400) // 2
        add_win.geometry(f"550x400+{x}+{y}")
        
        # Основной контейнер с отступами
        main_frame = ttk.Frame(add_win, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # --- Линия (Combobox) ---
        line_frame = ttk.Frame(main_frame)
        line_frame.pack(fill="x", pady=(0, 12))
        ttk.Label(line_frame, text="Линия:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        line_var = tk.StringVar(value=line_names[0] if line_names else "")
        line_combo = ttk.Combobox(line_frame, textvariable=line_var, values=line_names, font=("Segoe UI", 10), state="readonly")
        line_combo.pack(fill="x", pady=(2, 0))
        
        # --- Время и Тип в одной строке ---
        row_frame = ttk.Frame(main_frame)
        row_frame.pack(fill="x", pady=(0, 12))
        
        # Время
        time_frame = ttk.Frame(row_frame)
        time_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Label(time_frame, text="Время простоя:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        time_input_frame = ttk.Frame(time_frame)
        time_input_frame.pack(anchor="w", pady=(2, 0))
        time_var = tk.StringVar(value="10")
        time_entry = ttk.Entry(time_input_frame, textvariable=time_var, width=10, font=("Segoe UI", 10))
        time_entry.pack(side="left")
        ttk.Label(time_input_frame, text="мин", foreground="#666").pack(side="left", padx=(5, 0))
        
        # Тип (Combobox)
        type_frame = ttk.Frame(row_frame)
        type_frame.pack(side="left", fill="x", expand=True)
        ttk.Label(type_frame, text="Тип:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        type_var = tk.StringVar(value=type_options[0])
        type_combo = ttk.Combobox(type_frame, textvariable=type_var, values=type_options, width=20, font=("Segoe UI", 10))
        type_combo.pack(anchor="w", pady=(2, 0))
        
        # --- Описание ---
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(fill="x", pady=(0, 12))
        ttk.Label(desc_frame, text="Описание:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        desc_var = tk.StringVar()
        desc_entry = ttk.Entry(desc_frame, textvariable=desc_var, font=("Segoe UI", 10))
        desc_entry.pack(fill="x", pady=(2, 0))
        
        # --- Комментарий (многострочный) ---
        comment_frame = ttk.Frame(main_frame)
        comment_frame.pack(fill="both", expand=True, pady=(0, 15))
        ttk.Label(comment_frame, text="Комментарий:", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        comment_text = tk.Text(comment_frame, height=4, font=("Segoe UI", 10), wrap="word", relief="solid", borderwidth=1)
        comment_text.pack(fill="both", expand=True, pady=(2, 0))
        
        # --- Кнопки ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x")
        
        def add_new():
            line_name = line_var.get()
            try:
                time_val = float(time_var.get().replace(",", "."))
                if time_val <= 0:
                    raise ValueError()
            except:
                messagebox.showwarning("Ошибка", "Введите корректное время (положительное число)")
                return
            type_val = type_var.get()
            desc_val = desc_var.get()
            comment_val = comment_text.get("1.0", "end-1c").strip()
            
            if not desc_val.strip():
                messagebox.showwarning("Ошибка", "Введите описание проблемы")
                return
            
            # Добавляем в данные
            for ld in self.lines_data:
                if ld.line_name == line_name:
                    new_problem = ProblemEntry(ld.filepath, ld.sheet_name, "РУЧН", time_val, type_val, desc_val, comment_val)
                    ld.problems.append(new_problem)
                    ld.problems.sort(key=lambda x: x.time_val, reverse=True)
                    break
            
            self._populate_tree()
            self._refresh_summary()
            add_win.destroy()
        
        ttk.Button(btn_frame, text="Отмена", command=add_win.destroy, width=12).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Добавить", command=add_new, width=12, style="Primary.TButton").pack(side="right")

    def _delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Выберите", "Выберите проблему для удаления")
            return
        
        if not messagebox.askyesno("Подтверждение", "Удалить выбранную проблему?"):
            return
        
        item = selected[0]
        values = self.tree.item(item, "values")
        line_name, time_val, type_val, desc, comment = values
        
        # Удаляем из данных
        for ld in self.lines_data:
            if ld.line_name == line_name:
                ld.problems = [p for p in ld.problems if not (f"{p.time_val:g}" == time_val and p.formulation == desc)]
                break
        
        self.tree.delete(item)
        self._refresh_summary()

    def copy_text(self):
        text = self.txt_area.get("1.0", tk.END)
        self.clipboard_clear()
        self.clipboard_append(text)
        self.save_to_history()
        messagebox.showinfo("Готово", "Текст скопирован!")

    def save_to_history(self):
        if not self.lines_data:
            messagebox.showwarning("Пусто", "Нет данных для сохранения.")
            return
        d, m, y = self.app.get_selected_date()
        success, msg = self.app.history_manager.save_problems(self.lines_data, d, m, y)
        if success:
            messagebox.showinfo("История", msg)
        else:
            messagebox.showerror("История", msg)

    def create_consolidated_excel(self):
        if not self.lines_data:
            messagebox.showwarning("Пусто", "Нет данных для сохранения.")
            return

        default_name = f"Сводный_Отчет_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        reports_dir = os.path.abspath("Отчеты по сменам")
        try:
            os.makedirs(reports_dir, exist_ok=True)
        except Exception as e:
            messagebox.showwarning("Внимание", f"Не удалось создать папку отчетов:\n{e}")
            reports_dir = None

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            initialdir=reports_dir
        )
        if not save_path:
            return

        # Показываем прогресс и блокируем кнопку
        self.btn_excel.config(state="disabled")
        self.progress.pack(fill="x", padx=5, pady=(0, 5))
        self.progress["value"] = 0
        
        # Запускаем сборку в отдельном потоке
        threading.Thread(target=self._worker_consolidate, args=(save_path,), daemon=True).start()

    def _update_progress(self, value):
        """Обновление прогресса из GUI потока"""
        self.progress["value"] = value
    
    def _on_consolidate_done(self, success, message, save_path=None):
        """Callback завершения - вызывается в GUI потоке"""
        self.progress.pack_forget()
        self.btn_excel.config(state="normal")
        
        if success:
            messagebox.showinfo("Успех", message)
        else:
            messagebox.showerror("Ошибка сохранения", message)

    def _worker_consolidate(self, save_path):
        try:
            # Callback для прогресса (вызывается из worker thread)
            def on_progress(value):
                self.after(0, self._update_progress, value)

            def on_log(msg):
                print(msg)

            # Копирование через COM с fallback на openpyxl
            sheets_count, mode_label = copy_sheets_with_fallback(
                self.lines_data,
                save_path,
                progress_callback=on_progress,
                log_callback=on_log
            )

            self.after(
                0,
                self._on_consolidate_done,
                True,
                f"Файл собран!\nЛистов: {sheets_count}\nРежим: {mode_label}",
                save_path
            )

        except Exception as e:
            self.after(0, self._on_consolidate_done, False, str(e))

# --- GUI: ИСТОРИЯ ---
class HistoryTab(ttk.Frame):
    def __init__(self, parent, history_manager):
        super().__init__(parent)
        self.history_manager = history_manager
        today = datetime.now()
        self.current_month = today.month
        self.current_year = today.year
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        paned = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        self.left_frame = ttk.Frame(paned)
        self.right_frame = ttk.Frame(paned)
        paned.add(self.left_frame, weight=1)
        paned.add(self.right_frame, weight=2)

        # Навигация календаря
        nav = ttk.Frame(self.left_frame)
        nav.pack(fill="x", pady=(0, 10))
        ttk.Button(nav, text="<", width=3, command=self.prev_month).pack(side="left")
        self.month_label = ttk.Label(nav, text="", font=("Segoe UI", 10, "bold"))
        self.month_label.pack(side="left", expand=True)
        ttk.Button(nav, text=">", width=3, command=self.next_month).pack(side="right")

        # Заголовки дней недели
        header = ttk.Frame(self.left_frame)
        header.pack(fill="x")
        for i, day_name in enumerate(["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]):
            ttk.Label(header, text=day_name, width=4, anchor="center").grid(row=0, column=i, padx=2, pady=2)

        # Сетка календаря
        self.calendar_frame = ttk.Frame(self.left_frame)
        self.calendar_frame.pack(fill="both", expand=True)
        self.day_buttons = []
        for r in range(6):
            for c in range(7):
                btn = tk.Button(self.calendar_frame, text="", width=6, height=3, relief="flat", bg="#f5f5f5")
                btn.grid(row=r, column=c, padx=2, pady=2, sticky="nsew")
                self.day_buttons.append(btn)

        # Детали дня (справа)
        self.details_title = ttk.Label(self.right_frame, text="Проблемы за дату", font=("Segoe UI", 11, "bold"))
        self.details_title.pack(anchor="w", pady=(0, 10))

        columns = ("line", "time", "type", "description")
        self.details_tree = ttk.Treeview(self.right_frame, columns=columns, show="headings", height=15)
        self.details_tree.heading("line", text="Линия")
        self.details_tree.heading("time", text="Минуты")
        self.details_tree.heading("type", text="Тип")
        self.details_tree.heading("description", text="Описание")
        self.details_tree.column("line", width=150)
        self.details_tree.column("time", width=70)
        self.details_tree.column("type", width=120)
        self.details_tree.column("description", width=300)

        tree_scroll = ttk.Scrollbar(self.right_frame, orient="vertical", command=self.details_tree.yview)
        self.details_tree.configure(yscrollcommand=tree_scroll.set)
        self.details_tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        self.default_btn_bg = self.day_buttons[0].cget("bg")

    def refresh(self):
        self._render_calendar()

    def prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.refresh()

    def next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.refresh()

    def _render_calendar(self):
        month_name = MONTHS[self.current_month - 1]
        self.month_label.config(text=f"{month_name} {self.current_year}")

        stats = self.history_manager.get_month_stats(self.current_month, self.current_year)
        cal = calendar.Calendar(firstweekday=0)
        weeks = cal.monthdayscalendar(self.current_year, self.current_month)
        while len(weeks) < 6:
            weeks.append([0] * 7)

        btn_index = 0
        for week in weeks:
            for day in week:
                btn = self.day_buttons[btn_index]
                btn_index += 1

                if day == 0:
                    btn.config(text="", state="disabled", bg=self.default_btn_bg, command=lambda: None)
                    continue

                total = stats.get(day)
                if total is None or total == 0:
                    color = self.default_btn_bg
                elif total < 60:
                    color = "#d9ead3"
                elif total <= 180:
                    color = "#fff2cc"
                else:
                    color = "#f4cccc"

                total_text = f"{int(total)} мин" if total is not None else ""
                btn.config(
                    text=f"{day}\n({total_text})" if total_text else f"{day}",
                    state="normal",
                    bg=color,
                    command=lambda d=day: self.show_day(d)
                )

    def show_day(self, day):
        month_name = MONTHS[self.current_month - 1]
        self.details_title.config(text=f"Проблемы за {day} {month_name} {self.current_year}")

        for item in self.details_tree.get_children():
            self.details_tree.delete(item)

        rows = self.history_manager.get_day_details(day, self.current_month, self.current_year)
        for line_name, time_val, problem_type, description, comment in rows:
            desc = description or ""
            if comment:
                desc = f"{desc} | {comment}" if desc else comment
            self.details_tree.insert("", "end", values=(line_name, f"{time_val:g}", problem_type, desc))

# --- ОКНО ВЫБОРА ФАЙЛОВ ---
class FileSelectionWindow(tk.Toplevel):
    def __init__(self, parent, path_vars):
        super().__init__(parent)
        self.title("Файлы линий")
        self.geometry("800x600")
        self.path_vars = path_vars
        
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_inner = ttk.Frame(canvas)
        
        scroll_inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Прокрутка колесом
        self.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        for i in range(11):
            f = ttk.Frame(scroll_inner)
            f.pack(fill="x", pady=2)
            ttk.Label(f, text=f"{i+1}.", width=3).pack(side="left")
            ttk.Entry(f, textvariable=self.path_vars[i]).pack(side="left", fill="x", expand=True, padx=5)
            ttk.Button(f, text="...", width=3, command=lambda idx=i: self.browse(idx)).pack(side="left")
            
        ttk.Button(main_frame, text="Закрыть", command=self.destroy).pack(pady=10)

    def browse(self, idx):
        files = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if files:
            for i, p in enumerate(files):
                if idx + i < 11: self.path_vars[idx+i].set(os.path.normpath(p))

# --- ГЛАВНОЕ ПРИЛОЖЕНИЕ ---
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Reports & Consolidator v7.0")
        self.geometry("900x700")
        style = ttk.Style()
        style.theme_use('clam')
        self._setup_cyrillic()

        cfg = ConfigManager.load_config()
        self.path_vars = [tk.StringVar(value=p) for p in cfg["file_paths"]]
        self.min_time = tk.StringVar(value=str(cfg.get("min_downtime", 10)))
        self.excluded = tk.StringVar(value=cfg.get("excluded_categories", ""))
        self.history_manager = HistoryManager()
        
        # Дата - вчера
        yesterday = datetime.now() - timedelta(days=1)
        self.day = tk.StringVar(value=str(yesterday.day))
        self.month = tk.StringVar(value=MONTHS[yesterday.month - 1])
        self.year = tk.StringVar(value=str(yesterday.year))

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        
        self.tab_run = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_run, text="1. Запуск")
        
        self.tab_settings = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_settings, text="2. Настройки")
        
        self.tab_result = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_result, text="3. Результат и Сводка")

        self.tab_history = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_history, text="4. История")

        self._build_run()
        self._build_settings()
        self._build_result()
        self._build_history()
        self.cancel_event = threading.Event()

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _setup_cyrillic(self):
        self.bind_all("<Control-Cyrillic_es>", lambda e: e.widget.event_generate("<<Copy>>"))
        self.bind_all("<Control-Cyrillic_a>", lambda e: e.widget.event_generate("<<SelectAll>>"))

    def _build_run(self):
        frame = ttk.Frame(self.tab_run, padding=20)
        frame.pack(fill="both", expand=True)
        
        # Дата
        d_frame = ttk.LabelFrame(frame, text="Дата смены (Вчера)", padding=10)
        d_frame.pack(fill="x", pady=(0, 15))
        ttk.Combobox(d_frame, textvariable=self.day, values=[str(i) for i in range(1,32)], width=3).pack(side="left", padx=5)
        ttk.Combobox(d_frame, textvariable=self.month, values=MONTHS, width=10).pack(side="left", padx=5)
        ttk.Entry(d_frame, textvariable=self.year, width=6).pack(side="left", padx=5)

        # Кнопки
        style = ttk.Style()
        style.configure("Big.TButton", font=("Segoe UI", 14, "bold"), foreground="blue")
        self.btn_run = ttk.Button(frame, text="🚀 НАЧАТЬ АНАЛИЗ", command=self.run_process, style="Big.TButton")
        self.btn_run.pack(fill="x", ipady=15, pady=(0, 5))
        self.btn_cancel = ttk.Button(frame, text="⛔ ОТМЕНА", command=self.cancel_process, state="disabled")
        self.btn_cancel.pack(fill="x", ipady=5, pady=(0, 15))
        
        # Прогресс
        self.progress = ttk.Progressbar(frame, mode="determinate")
        self.progress.pack(fill="x", pady=(0, 10))

        # Сетка статусов
        self.status_grid = StatusGrid(frame, rows=11)
        self.status_grid.pack(fill="x", pady=(0, 10))
        
        # Лог
        self.log_widget = scrolledtext.ScrolledText(frame, height=10)
        self.log_widget.pack(fill="both", expand=True)

    def _build_settings(self):
        frame = ttk.Frame(self.tab_settings, padding=20)
        frame.pack(fill="both", expand=True)
        
        f_filt = ttk.LabelFrame(frame, text="Фильтры", padding=10)
        f_filt.pack(fill="x", pady=(0, 10))
        ttk.Label(f_filt, text="Мин. простой:").pack(side="left")
        ttk.Entry(f_filt, textvariable=self.min_time, width=5).pack(side="left", padx=10)
        ttk.Label(f_filt, text="Исключить:").pack(side="left")
        ttk.Entry(f_filt, textvariable=self.excluded).pack(side="left", fill="x", expand=True)
        
        ttk.Button(frame, text="📂 Список файлов", command=lambda: FileSelectionWindow(self, self.path_vars)).pack(fill="x", ipady=5)

    def _build_result(self):
        self.summary_frame = SummaryFrame(self.tab_result, self)
        self.summary_frame.pack(fill="both", expand=True)

    def _build_history(self):
        self.history_tab = HistoryTab(self.tab_history, self.history_manager)
        self.history_tab.pack(fill="both", expand=True)

    def log(self, msg):
        self.log_widget.insert(tk.END, msg + "\n")
        self.log_widget.see(tk.END)

    def get_selected_date(self):
        try:
            d = int(self.day.get())
        except:
            d = 1
        m = self.month.get()
        try:
            y = int(self.year.get())
        except:
            y = datetime.now().year
        return d, m, y

    def run_process(self):
        paths = [v.get() for v in self.path_vars]
        try:
            mt = int(self.min_time.get())
        except: mt = 0
        ConfigManager.save_config(paths, mt, self.excluded.get())
        
        self.log_widget.delete("1.0", tk.END)
        self.btn_run.config(state="disabled")
        self.btn_cancel.config(state="normal")
        self.cancel_event = threading.Event()
        self._init_status_grid(paths)
        
        try:
            d, m, y = self.get_selected_date()
        except: return

        threading.Thread(target=self._worker, args=(paths, d, m, y, mt, self.excluded.get()), daemon=True).start()

    def _worker(self, paths, d, m, y, mt, exc):
        proc = ExcelProcessor(
            lambda msg: self.after(0, self.log, msg),
            lambda v: self.after(0, lambda: self.progress.config(value=v)),
            lambda line, prog, status, message: self.after(0, self.status_grid.update_line, line, prog, status, message),
            self.cancel_event
        )
        data = proc.run(paths, d, m, y, mt, exc)
        summary_text = build_summary_text(data)
        self.after(0, self.finish, data, summary_text)

    def finish(self, data, summary_text=None):
        self.btn_run.config(state="normal")
        self.btn_cancel.config(state="disabled")
        self.summary_frame.populate(data, summary_text)
        if self.cancel_event.is_set():
            messagebox.showinfo("Отменено", f"Обработано линий: {len(data)}")
            return
        if data:
            messagebox.showinfo("Готово", f"Обработано линий: {len(data)}")
            self.notebook.select(self.tab_result)
        else:
            messagebox.showinfo("Пусто", "Нет данных по выбранной дате.")

    def cancel_process(self):
        self.cancel_event.set()
        self.log("Отмена обработки...")

    def _init_status_grid(self, paths):
        line_names = []
        for i, p in enumerate(paths):
            if p.strip():
                line_names.append(os.path.splitext(os.path.basename(p))[0])
            else:
                line_names.append(f"Линия {i+1}")
        self.status_grid.set_lines(line_names)

    def _on_tab_changed(self, event):
        if self.notebook.select() == str(self.tab_history):
            self.history_tab.refresh()

if __name__ == "__main__":
    app = App()
    app.mainloop()